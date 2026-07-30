"""
JSX Compliance Data Relay
=========================
Reads JSX Debriefs.xlsx on SharePoint (same Power Flows/Debriefs pattern as
the Envoy/PSA/Mesa/GoJet trackers), computes last-service dates and
compliance windows from the raw debrief rows, and writes data.json to the
repo root for the GitHub Pages dashboard to consume.

Runs on GitHub Actions on an hourly schedule.

Credentials required (set as GitHub Secrets):
  TENANT_ID       - Azure AD tenant ID
  CLIENT_ID       - Foxtrot Report Automation app ID
  CLIENT_SECRET   - Foxtrot Report Automation client secret

File locations:
  SharePoint Drive ID : b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU
  File path           : Power Flows/Debriefs/JSX Debriefs.xlsx
  Sheets read         : Sheet1 (Debriefs table), Sheet2 (tail roster)

Sheet2's "Last ..." columns are Excel array formulas whose cached values go
stale (Power Automate never recalculates them), so this script ignores them
and recomputes everything from the Sheet1 debrief rows.
"""

import os
import json
import sys
import requests
from datetime import datetime, timezone, date

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/JSX Debriefs.xlsx"

# Job cycle lengths in days
CYCLES = {"IC": 30, "EC": 90, "DSC": 30, "CE": 90}

# ─────────────────────────────────────────────
# STEP 1: Get Graph API token
# ─────────────────────────────────────────────

def get_token():
    print("Acquiring Graph API token...")
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    token = resp.json()["access_token"]
    print("  Token acquired.")
    return token


# ─────────────────────────────────────────────
# STEP 2: Download Excel file from SharePoint
# ─────────────────────────────────────────────

def download_excel(token):
    print(f"Downloading: {FILE_PATH}")
    encoded = FILE_PATH.replace(" ", "%20")
    url = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
        f"/root:/{encoded}:/content"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    path = "/tmp/jsx_debriefs.xlsx"
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Downloaded {len(resp.content):,} bytes → {path}")
    return path


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def flag(v):
    """0 for No/blank/0, 1 for anything else (handles 1/0, Yes/No, 'Yes. x')."""
    if v is None:
        return 0
    s = str(v).strip().lower()
    return 0 if s in ("no", "0", "", "none", "nan", "false") else 1


def parse_date_val(v):
    """Return a date or None. Handles datetime/date, ISO strings, Excel serials."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date (Power Automate sometimes writes these)
        from datetime import timedelta
        return date(1899, 12, 30) + timedelta(days=int(v))
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s[:10]).date()
    except ValueError:
        return None


# ─────────────────────────────────────────────
# STEP 3: Read tail roster — Sheet2
# ─────────────────────────────────────────────

def read_tail_roster(wb):
    """
    Sheet2 columns (0-indexed):
      0: Tail Number
      1: Plane Type
      2-6: Last Service / Last IC / Last EC / Last DSC / Last CE
           (array formulas — cached values are stale, ignored here)
      7: Status — "Disabled" hides the tail from the tracker; anything
         else (Active, blank) shows it, so a forgotten status on a newly
         added tail doesn't silently drop it.
    """
    ws = wb["Sheet2"]
    roster = []
    seen = set()
    disabled = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        tail = str(row[0] or "").strip().upper()
        if not tail or tail in seen:
            continue
        seen.add(tail)
        status = str(row[7] or "").strip().lower() if len(row) > 7 else ""
        if status == "disabled":
            disabled += 1
            continue
        roster.append({"tail": tail, "type": str(row[1] or "").strip()})
    print(f"  Tail roster (Sheet2): {len(roster)} active tails ({disabled} disabled)")
    return roster


# ─────────────────────────────────────────────
# STEP 4: Parse debriefs — Sheet1
# ─────────────────────────────────────────────

def parse_debriefs(wb):
    """
    Reads the Debriefs table on Sheet1. Columns expected (0-indexed):
      0:  Tail Number
      1:  Plane Type
      2:  Service Location
      3:  Date
      4:  Technician
      5:  Turn               (info only)
      6:  RON                (info only)
      7:  Interior Detail    (IC)
      8:  Exterior Detail    (EC)
      9:  Deep Seat Clean    (DSC)
      10: Carpet Extraction  (CE)
      11: Biohazard
      12: Sub ID
      13: Job Revenue        (not used by the dashboard)
    """
    ws = wb["Sheet1"]
    debriefs = []

    rows = list(ws.iter_rows(values_only=True))
    print(f"  Debriefs sheet: {len(rows)-1} data rows")

    for row in rows[1:]:
        tail = str(row[0] or "").strip().upper()
        if not tail:
            continue

        d = parse_date_val(row[3])

        debriefs.append({
            "tail":     tail,
            "type":     str(row[1] or ""),
            "location": str(row[2] or ""),
            "date":     d.isoformat() if d else None,
            "tech":     str(row[4] or ""),
            "turn":     flag(row[5]),
            "ron":      flag(row[6]),
            "IC":       flag(row[7]),
            "EC":       flag(row[8]),
            "DSC":      flag(row[9]),
            "CE":       flag(row[10]),
            "biohazard":flag(row[11]),
            "subId":    str(row[12] or ""),
            "link":     None,
        })

    print(f"  Parsed {len(debriefs)} debrief records.")
    return debriefs


# ─────────────────────────────────────────────
# STEP 4b: Build planes compliance table
# ─────────────────────────────────────────────

def build_planes(roster, debriefs):
    """
    For each tail in the roster, find the most recent date each tracked
    service (IC/EC/DSC/CE) was performed, then calculate compliance windows:
    window = cycle - days since last service ("No Service" if never).
    """
    today = date.today()

    by_tail = {}
    for d in debriefs:
        by_tail.setdefault(d["tail"], []).append(d)

    planes = []
    for entry in roster:
        tail = entry["tail"]
        records = by_tail.get(tail, [])

        last_dates = {job: None for job in CYCLES}
        last_service = None  # most recent debrief of any kind

        for d in records:
            if not d["date"]:
                continue
            d_date = datetime.strptime(d["date"], "%Y-%m-%d").date()
            if last_service is None or d_date > last_service:
                last_service = d_date
            for job in CYCLES:
                if d[job] == 1 and (last_dates[job] is None or d_date > last_dates[job]):
                    last_dates[job] = d_date

        windows = {}
        for job, cycle in CYCLES.items():
            if last_dates[job] is None:
                windows[job] = "No Service"
            else:
                windows[job] = cycle - (today - last_dates[job]).days

        planes.append({
            "tail":        tail,
            "type":        entry["type"],
            "lastService": last_service.isoformat() if last_service else None,
            "lastIC":      last_dates["IC"].isoformat()  if last_dates["IC"]  else None,
            "lastEC":      last_dates["EC"].isoformat()  if last_dates["EC"]  else None,
            "lastDSC":     last_dates["DSC"].isoformat() if last_dates["DSC"] else None,
            "lastCE":      last_dates["CE"].isoformat()  if last_dates["CE"]  else None,
            "IC":          windows["IC"],
            "EC":          windows["EC"],
            "DSC":         windows["DSC"],
            "CE":          windows["CE"],
        })

    print(f"  Built compliance for {len(planes)} planes.")
    return planes


# ─────────────────────────────────────────────
# STEP 5: Write data.json
# ─────────────────────────────────────────────

def write_json(planes, debriefs, requests=None):
    output = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "planes":    planes,
        "debriefs":  debriefs,
        "requests":  requests or [],
    }
    with open("data.json", "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"  Written: data.json  ({len(planes)} planes, {len(debriefs)} debriefs, {len(requests or [])} active requests)")




# ─────────────────────────────────────────────
# STEP 6: Process service requests
# ─────────────────────────────────────────────

def get_graph_token():
    """Get Microsoft Graph API token for sending emails."""
    import requests as req_lib
    r = req_lib.post(
        f"https://login.microsoftonline.com/{os.environ['TENANT_ID']}/oauth2/v2.0/token",
        data={
            "grant_type":    "client_credentials",
            "client_id":     os.environ["CLIENT_ID"],
            "client_secret": os.environ["CLIENT_SECRET"],
            "scope":         "https://graph.microsoft.com/.default",
        }
    )
    r.raise_for_status()
    return r.json()["access_token"]


def send_email(token, to_addresses, subject, body_html):
    """Send email via Graph API from foxtrot.automation."""
    import requests as req_lib
    recipients = [{"emailAddress": {"address": a}} for a in to_addresses]
    r = req_lib.post(
        "https://graph.microsoft.com/v1.0/users/foxtrot.automation@foxtrotaviation.com/sendMail",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={
            "message": {
                "subject": subject,
                "body": {"contentType": "HTML", "content": body_html},
                "toRecipients": recipients,
            }
        }
    )
    return r.status_code


SVC_NAMES = {
    "IC": "Interior Detail",
    "EC": "Exterior Detail",
    "DSC": "Deep Seat Clean",
    "CE": "Carpet Extraction",
}

DISTRO = "jsx.requests@foxtrotaviation.com"


def process_requests(debriefs):
    """
    Reads requests.json, checks each open request for fulfillment
    against debriefs since the request date, and returns active requests.

    Fulfillment: all requested services must appear (flag=1) across
    debriefs for that tail on or after the requestDate.

    Also sends:
      - Fulfillment email to the requestor when all services are complete
      - Warning email to the distro if any request is due in 1 day
    """
    req_path = "requests.json"
    if not os.path.exists(req_path):
        print("  No requests.json found — skipping")
        return []

    with open(req_path) as f:
        data = json.load(f)

    all_reqs = data.get("requests", [])
    active = []
    changed = False

    # Get Graph token once — only if we have credentials
    token = None
    has_creds = all(k in os.environ for k in ("TENANT_ID", "CLIENT_ID", "CLIENT_SECRET"))
    if has_creds:
        try:
            token = get_graph_token()
        except Exception as e:
            print(f"  Warning: Could not get Graph token for emails: {e}")

    today = date.today()

    for req in all_reqs:
        if req.get("status") != "open":
            continue

        tail      = req["tail"]
        req_date  = req["requestDate"]
        services  = req["services"]
        req_id    = req["requestId"]

        # Find debriefs for this tail on or after requestDate
        relevant = [d for d in debriefs if d["tail"] == tail and (d["date"] or "") >= req_date]

        # Union of services performed
        done = set()
        for d in relevant:
            for svc in services:
                if d.get(svc) == 1:
                    done.add(svc)

        if all(s in done for s in services):
            # ── FULFILLED ──────────────────────────────────
            print(f"  Request {req_id} FULFILLED — {tail} {services}")
            req["status"] = "fulfilled"
            changed = True

            if token:
                try:
                    svcs_str = ", ".join(SVC_NAMES.get(s, s) for s in services)
                    to = [req["requestorEmail"]]
                    if req.get("additionalEmail"):
                        to.append(req["additionalEmail"])

                    body = f"""
<p>Good news — your service request for <strong>{tail}</strong> has been fulfilled.</p>
<table style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;margin-top:12px">
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#f9f9f9">Tail Number</td><td style="padding:6px 14px">{tail}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555">Services Completed</td><td style="padding:6px 14px">{svcs_str}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#f9f9f9">Originally Requested By</td><td style="padding:6px 14px;background:#f9f9f9">{req_date}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555">Request ID</td><td style="padding:6px 14px;color:#888;font-size:12px">{req_id}</td></tr>
</table>
<p style="margin-top:16px">All requested services have been recorded in the Foxtrot JSX Compliance Tracker. No further action is needed.</p>
<p style="margin-top:16px;color:#888;font-size:12px">— Foxtrot Aviation Services JSX Compliance Tracker</p>
"""
                    code = send_email(token, to, f"[JSX Request Fulfilled] {tail} — {svcs_str}", body)
                    print(f"  Fulfillment email sent to {to} (status {code})")
                except Exception as e:
                    print(f"  Fulfillment email error: {e}")

        else:
            # ── STILL OPEN — check 1-day warning ───────────
            try:
                req_dt = datetime.strptime(req_date, "%Y-%m-%d").date()
                days_until = (req_dt - today).days
            except Exception:
                days_until = None

            if days_until is not None and days_until == 1 and not req.get('warned'):
                # Due tomorrow — send warning to distro
                remaining = [s for s in services if s not in done]
                completed = [s for s in services if s in done]
                svcs_remaining = ", ".join(SVC_NAMES.get(s, s) for s in remaining)
                svcs_done      = ", ".join(SVC_NAMES.get(s, s) for s in completed) if completed else "None"

                if token:
                    try:
                        body = f"""
<p><strong>⚠ Warning:</strong> The following service request is due <strong>tomorrow ({req_date})</strong> and has not yet been fully completed.</p>
<table style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;margin-top:12px">
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#fff8f0">Tail Number</td><td style="padding:6px 14px;background:#fff8f0">{tail}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555">Services Still Needed</td><td style="padding:6px 14px;color:#D97706;font-weight:bold">{svcs_remaining}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#fff8f0">Services Completed</td><td style="padding:6px 14px;background:#fff8f0">{svcs_done}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555">Due Date</td><td style="padding:6px 14px">{req_date}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#fff8f0">Requested By</td><td style="padding:6px 14px;background:#fff8f0">{req["requestorName"]} ({req["requestorEmail"]})</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555">Location</td><td style="padding:6px 14px">{req.get("location") or "Not specified"}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#555;background:#fff8f0">Notes</td><td style="padding:6px 14px;background:#fff8f0">{req.get("notes") or "—"}</td></tr>
</table>
<p style="margin-top:16px;color:#888;font-size:12px">— Foxtrot Aviation Services JSX Compliance Tracker</p>
"""
                        code = send_email(token, [DISTRO],
                            f"[JSX ⚠ Due Tomorrow] {tail} — {svcs_remaining}", body)
                        print(f"  1-day warning email sent for {req_id} (status {code})")
                        req['warned'] = True
                        changed = True
                    except Exception as e:
                        print(f"  Warning email error: {e}")

            active.append(req)

    # Write back if any requests were fulfilled
    if changed:
        data["requests"] = all_reqs
        with open(req_path, "w") as f:
            json.dump(data, f, indent=2)
        print("  requests.json updated with fulfilled statuses")

    print(f"  Active requests: {len(active)} of {len(all_reqs)} total")
    return active



# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl requests")
        sys.exit(1)

    print("=== JSX Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    token     = get_token()
    xlsx_path = download_excel(token)

    print("\nParsing workbook...")
    wb       = openpyxl.load_workbook(xlsx_path, data_only=True)
    roster   = read_tail_roster(wb)
    debriefs = parse_debriefs(wb)

    print("\nBuilding compliance table...")
    planes   = build_planes(roster, debriefs)

    print("\nWriting data.json...")
    print("\nProcessing service requests...")
    active_requests = process_requests(debriefs)
    write_json(planes, debriefs, active_requests)

    print("\n=== Done ===")
